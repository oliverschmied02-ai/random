"""
Generate reports from the listing database:
  - listings_master.csv
  - listings_report.xlsx  (color-coded by attractiveness score)
  - dashboard.html        (interactive table + Folium map)
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from src.config import AppConfig
from src.models.listing import Listing

logger = logging.getLogger(__name__)


def generate_reports(listings: list[Listing], config: AppConfig) -> dict[str, Path]:
    """Generate all configured report formats. Returns {format: path}."""
    output_dir = Path(config.reporting.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    formats = config.reporting.output_formats
    generated: dict[str, Path] = {}

    if not listings:
        logger.info("No listings to report on")
        return generated

    if "csv" in formats:
        path = _write_csv(listings, output_dir)
        generated["csv"] = path

    if "excel" in formats:
        path = _write_excel(listings, output_dir)
        if path:
            generated["excel"] = path

    if "html" in formats:
        path = _write_html(listings, output_dir)
        if path:
            generated["html"] = path

    logger.info("Reports written to %s: %s", output_dir, list(generated.keys()))
    return generated


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def _write_csv(listings: list[Listing], output_dir: Path) -> Path:
    import csv
    path = output_dir / "listings_master.csv"
    rows = [l.to_dict() for l in listings]
    if not rows:
        return path
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    logger.info("CSV: %s (%d rows)", path, len(rows))
    return path


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _write_excel(listings: list[Listing], output_dir: Path) -> Optional[Path]:
    try:
        import pandas as pd
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("pandas/openpyxl not installed — skipping Excel report")
        return None

    path = output_dir / f"zvg_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    rows = [l.to_dict() for l in listings]
    df = pd.DataFrame(rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Listings")
        wb = writer.book
        ws = writer.sheets["Listings"]

        # Header styling
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Color-code rows by attractiveness score
        for row_idx, listing in enumerate(listings, start=2):
            score = listing.ai_attractiveness_score
            if score is None:
                continue
            fill_color = _score_color(score)
            fill = PatternFill("solid", fgColor=fill_color)
            for cell in ws[row_idx]:
                cell.fill = fill

        # Auto-fit column widths (capped at 50)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=10
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    logger.info("Excel: %s", path)
    return path


def _score_color(score: int) -> str:
    """Return Excel hex fill color for attractiveness score."""
    if score >= 8:
        return "C6EFCE"  # green
    if score >= 6:
        return "FFEB9C"  # yellow
    if score >= 4:
        return "FFCC99"  # orange
    return "FFC7CE"      # red


# ---------------------------------------------------------------------------
# HTML Dashboard
# ---------------------------------------------------------------------------

def _write_html(listings: list[Listing], output_dir: Path) -> Optional[Path]:
    path = output_dir / "dashboard.html"
    map_html = _build_map(listings)
    table_html = _build_table(listings)

    html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<title>ZVG Intelligence Dashboard</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 0; padding: 16px; background: #f5f5f5; }}
  h1 {{ color: #1F4E79; }}
  .stats {{ display: flex; gap: 16px; margin-bottom: 16px; }}
  .stat-card {{ background: white; padding: 12px 20px; border-radius: 8px;
                box-shadow: 0 1px 4px rgba(0,0,0,.1); min-width: 140px; }}
  .stat-card .value {{ font-size: 2em; font-weight: bold; color: #1F4E79; }}
  .stat-card .label {{ color: #666; font-size: .85em; }}
  #map {{ height: 450px; margin-bottom: 24px; border-radius: 8px;
          box-shadow: 0 1px 4px rgba(0,0,0,.1); }}
  table {{ width: 100%; border-collapse: collapse; background: white;
           border-radius: 8px; overflow: hidden;
           box-shadow: 0 1px 4px rgba(0,0,0,.1); }}
  th {{ background: #1F4E79; color: white; padding: 10px; text-align: left; cursor: pointer; }}
  td {{ padding: 8px 10px; border-bottom: 1px solid #eee; font-size: .9em; }}
  tr:hover td {{ background: #f0f7ff; }}
  .score-high {{ background: #C6EFCE; }}
  .score-mid  {{ background: #FFEB9C; }}
  .score-low  {{ background: #FFC7CE; }}
  .cancelled  {{ color: #999; text-decoration: line-through; }}
  input#search {{ width: 100%; padding: 10px; margin-bottom: 12px; border: 1px solid #ccc;
                  border-radius: 4px; font-size: 1em; }}
</style>
</head>
<body>
<h1>🏠 ZVG Intelligence Dashboard</h1>
<p>Stand: {datetime.utcnow().strftime("%d.%m.%Y %H:%M")} UTC</p>

{_stats_cards(listings)}

<div id="map">{map_html}</div>

<input type="text" id="search" placeholder="Filtern: Aktenzeichen, Ort, PLZ ..." onkeyup="filterTable()">

{table_html}

<script>
function filterTable() {{
  const q = document.getElementById('search').value.toLowerCase();
  document.querySelectorAll('#listings-table tbody tr').forEach(row => {{
    row.style.display = row.textContent.toLowerCase().includes(q) ? '' : 'none';
  }});
}}
// Simple click-to-sort
document.querySelectorAll('th[data-col]').forEach(th => {{
  th.addEventListener('click', () => {{
    const col = +th.dataset.col;
    const tbody = document.querySelector('#listings-table tbody');
    const rows = Array.from(tbody.rows);
    const asc = th.dataset.asc !== '1';
    th.dataset.asc = asc ? '1' : '0';
    rows.sort((a, b) => {{
      const va = a.cells[col]?.textContent.trim() ?? '';
      const vb = b.cells[col]?.textContent.trim() ?? '';
      const na = parseFloat(va.replace(/[^\\d.]/g, ''));
      const nb = parseFloat(vb.replace(/[^\\d.]/g, ''));
      if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
      return asc ? va.localeCompare(vb) : vb.localeCompare(va);
    }});
    rows.forEach(r => tbody.appendChild(r));
  }});
}});
</script>
</body>
</html>"""

    path.write_text(html, encoding="utf-8")
    logger.info("HTML dashboard: %s", path)
    return path


def _stats_cards(listings: list[Listing]) -> str:
    active = [l for l in listings if l.status == "active"]
    scored = [l for l in listings if l.ai_attractiveness_score is not None]
    avg_score = (
        sum(l.ai_attractiveness_score for l in scored) / len(scored) if scored else 0
    )
    avg_vw = (
        sum(l.verkehrswert for l in active if l.verkehrswert) /
        max(1, sum(1 for l in active if l.verkehrswert))
    )
    return f"""
<div class="stats">
  <div class="stat-card"><div class="value">{len(listings)}</div><div class="label">Gesamt</div></div>
  <div class="stat-card"><div class="value">{len(active)}</div><div class="label">Aktiv</div></div>
  <div class="stat-card"><div class="value">{avg_score:.1f}</div><div class="label">Ø Score</div></div>
  <div class="stat-card"><div class="value">{avg_vw/1000:.0f}k €</div><div class="label">Ø Verkehrswert</div></div>
</div>"""


def _build_map(listings: list[Listing]) -> str:
    try:
        import folium
    except ImportError:
        return "<p><em>Karte nicht verfügbar (folium nicht installiert)</em></p>"

    geo = [l for l in listings if l.lat and l.lon]
    if not geo:
        return "<p><em>Keine Geodaten verfügbar</em></p>"

    center_lat = sum(l.lat for l in geo) / len(geo)
    center_lon = sum(l.lon for l in geo) / len(geo)
    m = folium.Map(location=[center_lat, center_lon], zoom_start=9)

    for listing in geo:
        color = (
            "green" if (listing.ai_attractiveness_score or 0) >= 7
            else "orange" if (listing.ai_attractiveness_score or 0) >= 4
            else "red"
        )
        vw_str = f"{listing.verkehrswert:,.0f} €" if listing.verkehrswert is not None else "k.A."
        popup_html = (
            f"<b>{listing.aktenzeichen}</b><br>"
            f"{listing.amtsgericht}<br>"
            f"Verkehrswert: {vw_str}<br>"
            f"Score: {listing.ai_attractiveness_score or '–'}/10"
        )
        folium.CircleMarker(
            location=[listing.lat, listing.lon],
            radius=8,
            color=color,
            fill=True,
            popup=folium.Popup(popup_html, max_width=250),
            tooltip=listing.aktenzeichen,
        ).add_to(m)

    return m._repr_html_()


def _build_table(listings: list[Listing]) -> str:
    headers = [
        ("Aktenzeichen", 0), ("Amtsgericht", 1), ("Termin", 2),
        ("Ort", 3), ("Verkehrswert €", 4), ("Mindestgebot €", 5),
        ("Wohnfl. m²", 6), ("€/m²", 7), ("Score", 8), ("Risiken", 9),
    ]
    ths = "".join(
        f'<th data-col="{i}">{h}</th>' for h, i in headers
    )

    rows_html = ""
    for l in sorted(
        listings,
        key=lambda x: (x.ai_attractiveness_score or 0),
        reverse=True,
    ):
        score = l.ai_attractiveness_score
        cls = (
            "score-high" if score and score >= 7
            else "score-mid" if score and score >= 4
            else "score-low" if score
            else ""
        )
        if l.status == "cancelled":
            cls += " cancelled"

        drive_link = (
            f'<a href="{l.drive_folder_url}" target="_blank">Drive</a>'
            if l.drive_folder_url
            else ""
        )
        rows_html += f"""<tr class="{cls}">
  <td>{l.aktenzeichen} {drive_link}</td>
  <td>{l.amtsgericht}</td>
  <td>{l.termin.strftime("%d.%m.%Y") if l.termin else ""}</td>
  <td>{l.plz} {l.ort}</td>
  <td>{f"{l.verkehrswert:,.0f}" if l.verkehrswert else "–"}</td>
  <td>{f"{l.mindestgebot_50pct:,.0f}" if l.mindestgebot_50pct else "–"}</td>
  <td>{l.wohnflaeche_sqm or "–"}</td>
  <td>{f"{l.price_per_sqm:,.0f}" if l.price_per_sqm else "–"}</td>
  <td>{score or "–"}</td>
  <td>{"; ".join(l.ai_risk_flags[:3])}</td>
</tr>"""

    return f"""<table id="listings-table">
<thead><tr>{ths}</tr></thead>
<tbody>{rows_html}</tbody>
</table>"""
